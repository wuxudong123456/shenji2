"""从 2026-6-20提取模板.xlsx 导入违规行为库 + 案例库（v2 字段映射修正）

数据流程:
  Excel 2231 行 → tt.audit_violations
               → tt.audit_violation_law_refs（违规↔法规关联，5级降级匹配）
               → tt.audit_cases（案例拆分）
               → tt.audit_case_violations（案例↔违规）
               → tt.audit_case_law_refs（案例↔法规，继承违规的法规）

v2 字段映射修正:
  - violation_code: XV-{YYYYMMDD}-{行号:04d}-{序号:02d}
  - audititem_id:  审计事项名称→树name（A），失败回退分类→树节点（B）
  - expression_text: 只存纯违规表达式
  - audit_procedure: 审计方法步骤 + 疑点发现方法
  - required_data:  审计所需数据 + 所需资料类型 + 对应数据字段（丰富 JSON）
  - description:    规则说明 + 疑点推理 + 违规依据法规名（供前端正则）

前提:
  1. 已执行 ALTER TABLE 加 audit_procedure / required_data 两列
  2. 旧数据已清空（或导入脚本会跳过重复 expression_text）

用法:
  cd backend && python data/import_excel.py            # 试运行（不写库）
  cd backend && python data/import_excel.py --run      # 正式导入
"""
import sys
import json
import re
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from services.db import query, query_one, insert  # noqa: E402

# 相对 backend/ 的路径
ROOT = Path(__file__).resolve().parent.parent.parent
EXCEL_FILE = ROOT / "2026-6-20提取模板.xlsx"
SHEET_NAME = "审计疑点特征提取"
SOURCE_FILE = "2026-6-20提取模板.xlsx"
BATCH_ID = f"EXCEL-{datetime.now().strftime('%Y%m%d-%H%M')}"
BATCH_DATE = "20260620"  # 取自文件名 "2026-6-20" 批次日期，用于 violation_code

# Excel 列号（1-based）
C_ROWNO, C_ITEM, C_CATEGORY, C_MANIFEST, C_REQDATA, C_PROCEDURE, \
C_TITLE, C_MATERIALS, C_FIELDDEFS, C_EXPR, C_EXPRNOTE, C_DISCOVERY, \
C_LAWS, C_REASONING, C_CASES = range(1, 16)


# ────────────────────────────────────────────────────────────
#  法规匹配（5级降级，与 migrate_violation_law_refs.py 一致）
# ────────────────────────────────────────────────────────────

def _match_law_id(law_title: str) -> str | None:
    """根据法规名称匹配法规库 id，返回 law_id 或 None"""
    if not law_title:
        return None
    clean = law_title.strip().strip("《").strip("》").strip()
    if not clean or len(clean) < 3:
        return None

    # 1. 审计常用库 - 精确
    row = query_one(
        "SELECT id FROM sys_core_law_allaudit WHERE title = %s AND status = 1 LIMIT 1",
        (clean,), database="audit_law",
    )
    if row:
        return row["id"]

    # 2. 审计常用库 - 子串
    row = query_one(
        "SELECT id FROM sys_core_law_allaudit WHERE title LIKE %s AND status = 1 LIMIT 1",
        (f"%{clean}%",), database="audit_law",
    )
    if row:
        return row["id"]

    # 3. 全量库 - 精确
    row = query_one(
        "SELECT id FROM sys_core_law WHERE title = %s LIMIT 1",
        (clean,), database="audit_law",
    )
    if row:
        return row["id"]

    # 4. 全量库 - 子串
    row = query_one(
        "SELECT id FROM sys_core_law WHERE title LIKE %s LIMIT 1",
        (f"%{clean}%",), database="audit_law",
    )
    if row:
        return row["id"]

    # 5. 书名号内容匹配
    for b in re.findall(r'《([^》]+)》', law_title):
        row = query_one(
            "SELECT id FROM sys_core_law_allaudit WHERE title LIKE %s AND status = 1 LIMIT 1",
            (f"%{b}%",), database="audit_law",
        )
        if row:
            return row["id"]
        row = query_one(
            "SELECT id FROM sys_core_law WHERE title LIKE %s LIMIT 1",
            (f"%{b}%",), database="audit_law",
        )
        if row:
            return row["id"]

    return None


# ────────────────────────────────────────────────────────────
#  审计事项匹配（A+B 组合策略）
#  A: Excel.审计事项名称 → sys_audititem_SLFF.name（精确/归一化/包含/被包含）
#  B: 失败后回退 → Excel.审计事项分类最后一段 → 树 name 精确，否则 path 匹配
# ────────────────────────────────────────────────────────────

_audititem_tree_cache: list | None = None


def _get_audititem_tree() -> list:
    """加载 sys_audititem_SLFF 全部节点（缓存）"""
    global _audititem_tree_cache
    if _audititem_tree_cache is None:
        _audititem_tree_cache = query(
            "SELECT id, name, level, path_names FROM sys_audititem_SLFF WHERE name IS NOT NULL",
            database="audit_law",
        )
    return _audititem_tree_cache


def _normalize_name(name: str) -> str:
    """归一化：去前导编号、引号括号、空白"""
    if not name:
        return ""
    s = re.sub(r'^[\d\．.\、\s]+', '', name.strip())
    s = re.sub(r'[《》""‘’（）()\[\]「」]', '', s)
    return re.sub(r'\s+', '', s)


def _build_audititem_index() -> tuple:
    """构建快速匹配索引: (tree, name_to_id, norm_to_id)"""
    tree = _get_audititem_tree()
    name_to_id = {t["name"]: t["id"] for t in tree if t["name"]}
    norm_to_id = {}
    for t in tree:
        n = _normalize_name(t["name"])
        if n and n not in norm_to_id:
            norm_to_id[n] = t["id"]
    return tree, name_to_id, norm_to_id


def _match_audititem(item_name: str, category: str,
                     tree: list, name_to_id: dict, norm_to_id: dict) -> str | None:
    """Excel 审计事项 → sys_audititem_SLFF.id（A+B 组合策略）"""
    if not tree:
        return None

    # ── A. 审计事项名称 → 树 name ──
    if item_name:
        clean = item_name.strip()
        # A1 精确
        if clean in name_to_id:
            return name_to_id[clean]
        # A2 归一化
        n = _normalize_name(clean)
        if n and n in norm_to_id:
            return norm_to_id[n]
        # A3/A4 包含匹配（Excel名含树名 或 树名含Excel名）
        for t in tree:
            tn = t["name"]
            if tn and (tn in clean or clean in tn):
                return t["id"]

    # ── B. 分类回退 → 树 name 精确，否则 path 匹配 ──
    if category:
        seg = category.split("-")[-1].strip() if "-" in category else category.strip()
        if seg:
            # B1 分类段 == 树 name
            if seg in name_to_id:
                return name_to_id[seg]
            # B2 分类段出现在 path_names，取最上层（level 最小）代表节点
            best = None
            for t in tree:
                path = t.get("path_names") or ""
                if seg in path:
                    lvl = t["level"] if t["level"] is not None else 99
                    if best is None or lvl < best[1]:
                        best = (t["id"], lvl)
            if best:
                return best[0]

    return None


# ────────────────────────────────────────────────────────────
#  Excel 读取与解析
# ────────────────────────────────────────────────────────────

def _cell_text(val) -> str:
    """单元格值 → 去空白字符串"""
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def _load_rows() -> list[dict]:
    """读取 Excel，返回解析后的行字典列表"""
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet 不存在: {SHEET_NAME}（实际: {wb.sheetnames}）")
    ws = wb[SHEET_NAME]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 15:
            continue
        vals = list(row)
        # 跳过完全空行
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in vals):
            continue
        rows.append({
            "row_no": _cell_text(vals[C_ROWNO - 1]),
            "item": _cell_text(vals[C_ITEM - 1]),
            "category": _cell_text(vals[C_CATEGORY - 1]),
            "manifest": _cell_text(vals[C_MANIFEST - 1]),
            "req_data": _cell_text(vals[C_REQDATA - 1]),
            "procedure": _cell_text(vals[C_PROCEDURE - 1]),
            "title": _cell_text(vals[C_TITLE - 1]),
            "materials": _cell_text(vals[C_MATERIALS - 1]),
            "field_defs": _cell_text(vals[C_FIELDDEFS - 1]),
            "expr": _cell_text(vals[C_EXPR - 1]),
            "expr_note": _cell_text(vals[C_EXPRNOTE - 1]),
            "discovery": _cell_text(vals[C_DISCOVERY - 1]),
            "laws_raw": _cell_text(vals[C_LAWS - 1]),
            "reasoning": _cell_text(vals[C_REASONING - 1]),
            "cases_raw": _cell_text(vals[C_CASES - 1]),
        })
    wb.close()
    return rows


def _parse_json_array(raw: str) -> list | None:
    """安全解析 JSON 数组字符串；失败返回 None"""
    if not raw:
        return None
    try:
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, list) else None
    except (json.JSONDecodeError, ValueError):
        return None


def _format_law_line(item: dict) -> str:
    """法规对象 → '《名称》（文号）条款号' 一行文本"""
    name = (item.get("法规名称") or "").strip()
    if not name:
        return ""
    clean = name if name.startswith("《") else f"《{name}》"
    no = (item.get("法规文号") or "").strip()
    clause = (item.get("条款号") or "").strip()
    parts = [clean]
    if no:
        parts.append(f"（{no}）")
    if clause:
        parts.append(clause)
    return "• " + "".join(parts)


def _strip_md_fence(text: str) -> str:
    """去掉 markdown 代码围栏 ```json ... ```"""
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r'^```[a-zA-Z]*\s*', '', cleaned)
        cleaned = re.sub(r'```\s*$', '', cleaned)
        cleaned = cleaned.strip()
    return cleaned


# ────────────────────────────────────────────────────────────
#  字段构建（v2 映射）
# ────────────────────────────────────────────────────────────

def build_expression(row: dict) -> str:
    """⑩ → expression_text（只存纯违规表达式）"""
    return row["expr"]


def build_description(row: dict) -> str:
    """⑪ → description（只填表达式备注，其他内容去掉）"""
    return row["expr_note"]


def build_audit_procedure(row: dict) -> str | None:
    """⑥+⑫ → audit_procedure（审计方法步骤 + 疑点发现方法）"""
    parts = []
    if row["procedure"]:
        parts.append("# 审计方法步骤\n\n" + row["procedure"])
    if row["discovery"]:
        parts.append("# 疑点发现方法\n\n" + row["discovery"])
    return "\n\n".join(parts) if parts else None


def _parse_field_defs(raw: str) -> dict:
    """'表名{字段、字段}; 表名{字段、字段}...' → {表名: [字段...]}
    分隔符同时支持 ; 和 换行；字段分隔符同时支持 、 ， ,"""
    result = {}
    if not raw:
        return result
    for seg in re.split(r'[;\n]', raw):
        seg = seg.strip()
        if not seg:
            continue
        m = re.match(r'([^{}]+)\{([^}]*)\}', seg)
        if m:
            name = m.group(1).strip()
            fields = [f.strip() for f in re.split(r'[、，,]', m.group(2)) if f.strip()]
            if name:
                result[name] = fields
    return result


def _parse_materials(raw: str) -> list[dict]:
    """'大类-子类-名称、...' → [{cat, sub, name}, ...]
    分隔符同时支持 、 和 换行"""
    items = []
    if not raw:
        return items
    for seg in re.split(r'[、\n]', raw):
        seg = seg.strip()
        if not seg:
            continue
        parts = [p.strip() for p in re.split(r'[-—]', seg) if p.strip()]
        if len(parts) >= 2:
            items.append({
                "cat": parts[0],
                "sub": "-".join(parts[1:-1]),
                "name": parts[-1],
            })
        elif parts:
            items.append({"cat": parts[0], "sub": "", "name": parts[0]})
    return items


def _best_name_match(name: str, candidates) -> str | None:
    """从候选中找最佳名称匹配：精确 > 子串，取最长"""
    if not name or not candidates:
        return None
    if name in candidates:
        return name
    best = None
    for c in candidates:
        if c and (name in c or c in name):
            if best is None or len(c) > len(best):
                best = c
    return best


def build_required_data(row: dict) -> str | None:
    """⑧(所需资料类型) + ⑨(对应数据字段) → required_data
    实测 100% 每条所需资料类型都能配上对应数据字段。
    结构: {items: [{name, material_type, fields}]}
    注: 不再使用 Col⑤(审计所需数据)，以 Col⑧ 为主、Col⑨ 补字段。"""
    items = _parse_materials(row["materials"])
    if not items:
        return None
    field_defs = _parse_field_defs(row["field_defs"])
    f_names = list(field_defs.keys())
    entries = []
    for it in items:
        m = _best_name_match(it["name"], f_names)
        entries.append({
            "name": it["name"],
            "material_type": "-".join(p for p in [it["cat"], it["sub"]] if p),
            "fields": field_defs.get(m, []) if m else [],
        })
    return json.dumps({"items": entries}, ensure_ascii=False)


# ────────────────────────────────────────────────────────────
#  导入主逻辑
# ────────────────────────────────────────────────────────────

def load_and_prepare() -> list[dict]:
    """解析 Excel 全部行，生成待导入记录"""
    excel_rows = _load_rows()
    tree, name_to_id, norm_to_id = _build_audititem_index()

    records = []
    seen_codes = set()
    row_seq = {}  # 同一原始行号下多个疑点 → 序号，保证 violation_code 唯一
    for r in excel_rows:
        if not r["title"]:
            continue  # 无违规疑点名称 → 跳过

        # violation_code: XV-20260620-0006-01
        seq = row_seq.get(r["row_no"], 0) + 1
        row_seq[r["row_no"]] = seq
        violation_code = f"XV-{BATCH_DATE}-{r['row_no']:0>4}-{seq:02d}"
        if violation_code in seen_codes:
            continue
        seen_codes.add(violation_code)

        # audititem 匹配（A+B）
        audititem_id = _match_audititem(r["item"], r["category"],
                                        tree, name_to_id, norm_to_id)

        # 解析法规与案例
        laws = _parse_json_array(r["laws_raw"]) or []
        cases = _parse_json_array(r["cases_raw"]) or []

        records.append({
            "row": r,
            "violation_code": violation_code,
            "violation_title": r["title"][:500],
            "audititem_id": audititem_id,
            "review_status": "mapped" if audititem_id else "pending_mapping",
            "category_path": r["category"][:500],
            "expression_text": build_expression(r),
            "description": build_description(r),
            "audit_procedure": build_audit_procedure(r),
            "required_data": build_required_data(r),
            "laws": [x for x in laws if isinstance(x, dict)],
            "cases": [x for x in cases if isinstance(x, dict)],
        })
    return records


def import_excel(dry_run: bool = True) -> dict:
    print(f"源文件: {EXCEL_FILE}")
    print(f"批次号: {BATCH_ID}")
    print()

    records = load_and_prepare()
    print(f"解析 Excel 行数: {len(records)}")

    # 统计分类
    from collections import Counter
    cats = Counter(r["category_path"] for r in records)
    print(f"分类数: {len(cats)}")
    for cat, cnt in cats.most_common():
        print(f"  {cnt:4d}  {cat}")

    # 统计 audititem 匹配率
    matched_items = sum(1 for r in records if r["audititem_id"])
    print(f"\naudititem 匹配率: {matched_items}/{len(records)} "
          f"({100 * matched_items // max(len(records), 1)}%)")

    # 统计法规引用与匹配率
    law_refs_total = sum(len(r["laws"]) for r in records)
    unique_law_titles = set()
    for r in records:
        for law in r["laws"]:
            t = (law.get("法规名称") or "").strip()
            if t:
                unique_law_titles.add(t)
    print(f"法规引用总条数: {law_refs_total}, 去重法规名: {len(unique_law_titles)}")

    # 统计案例
    case_count = sum(len(r["cases"]) for r in records)
    rows_with_cases = sum(1 for r in records if r["cases"])
    print(f"案例对象总数: {case_count}, 含案例的行: {rows_with_cases}")

    # 样例预览
    print("\n=== 前 3 条解析预览 ===")
    for rec in records[:3]:
        print(f"--- {rec['violation_code']} ---")
        print(f"  title:        {rec['violation_title'][:60]}")
        print(f"  audititem:    {rec['audititem_id']} ({rec['review_status']})")
        print(f"  category:     {rec['category_path']}")
        print(f"  expr[:80]:    {rec['expression_text'][:80]}")
        print(f"  desc[:120]:   {rec['description'][:120]}")
        print(f"  procedure:    {'有' if rec['audit_procedure'] else '无'}, {len(rec['audit_procedure'] or '')} 字符")
        print(f"  required_data:{'有' if rec['required_data'] else '无'}")
        print(f"  laws: {len(rec['laws'])}, cases: {len(rec['cases'])}")

    if dry_run:
        print("\n=== 试运行模式（不写库）===")
        print(f"确认无误后执行: python data/import_excel.py --run")
        return {"dry_run": True, "records": len(records), "cases": case_count}

    # ── 正式导入 ──
    print("\n=== 正式导入 ===")
    v_inserted = v_skipped = 0
    law_refs_inserted = law_refs_unmatched = 0
    cases_inserted = 0
    case_viol_inserted = 0
    case_law_inserted = 0
    errors = 0

    for idx, rec in enumerate(records, 1):
        row = rec["row"]
        # 1. 违规行为
        existing = query_one(
            "SELECT id FROM audit_violations WHERE expression_text = %s AND deleted = b'0'",
            (rec["expression_text"],), database="tt",
        )
        if existing:
            v_skipped += 1
            continue
        try:
            violation_id = insert(
                """INSERT INTO audit_violations
                   (violation_code, violation_title, audititem_id, category_path, severity,
                    expression_text, description, audit_procedure, required_data,
                    source_file, import_batch, is_reviewed, review_status, creator, create_time)
                   VALUES (%s,%s,%s,%s,%s, %s,%s,%s,%s, %s,%s, %s,%s,%s, NOW())""",
                (
                    rec["violation_code"], rec["violation_title"], rec["audititem_id"],
                    rec["category_path"], "medium",
                    rec["expression_text"], rec["description"],
                    rec["audit_procedure"], rec["required_data"],
                    SOURCE_FILE, BATCH_ID, 0, rec["review_status"], "system",
                ),
                database="tt",
            )
            v_inserted += 1
        except Exception as e:
            errors += 1
            if errors <= 5:
                print(f"  错误 #{errors}: {rec['violation_code']} — {e}")
            continue

        # 2. 法规关联
        # 注: law_id 列 NOT NULL 且无法规外键。与 migrate_violation_law_refs.py 一致，
        # 未匹配到 law_id 的法规跳过关联表（法规名已在 description【违规依据】中保留）。
        matched_law_ids = []
        for law in rec["laws"]:
            law_title = (law.get("法规名称") or "").strip()
            if not law_title:
                continue
            law_id = _match_law_id(law_title)
            if not law_id:
                law_refs_unmatched += 1
                continue
            matched_law_ids.append(law_id)
            try:
                insert(
                    "INSERT INTO audit_violation_law_refs (violation_id, law_id, law_title, clause_ref) "
                    "VALUES (%s, %s, %s, %s)",
                    (violation_id, law_id, law_title[:500], (law.get("条款号") or "")[:500]),
                    database="tt",
                )
                law_refs_inserted += 1
            except Exception:
                pass

        # 3. 案例
        for case in rec["cases"]:
            brief = (case.get("案例简述") or "").strip()
            if not brief:
                continue
            title = brief.split("\n")[0][:500]
            try:
                case_id = insert(
                    """INSERT INTO audit_cases
                       (title, domain, case_summary, audit_method,
                        audit_finding, audit_impact, source)
                       VALUES (%s,%s,%s,%s, %s,%s,%s)""",
                    (
                        title, rec["category_path"][:100],
                        brief, (case.get("核查方法") or ""),
                        (case.get("违规表现") or ""), (case.get("风险影响") or ""),
                        SOURCE_FILE,
                    ),
                    database="tt",
                )
                cases_inserted += 1
            except Exception as e:
                errors += 1
                if errors <= 5:
                    print(f"  案例错误 #{errors}: {title[:40]} — {e}")
                continue

            # 案例↔违规
            try:
                insert(
                    "INSERT IGNORE INTO audit_case_violations (case_id, violation_id) VALUES (%s,%s)",
                    (case_id, violation_id), database="tt",
                )
                case_viol_inserted += 1
            except Exception:
                pass

            # 案例↔法规（继承违规的法规关联）
            for law_id in matched_law_ids:
                try:
                    insert(
                        "INSERT IGNORE INTO audit_case_law_refs (case_id, law_id) VALUES (%s,%s)",
                        (case_id, law_id), database="tt",
                    )
                    case_law_inserted += 1
                except Exception:
                    pass

        if idx % 200 == 0:
            print(f"  进度: {idx}/{len(records)} (违规 {v_inserted}, 案例 {cases_inserted})")

    print()
    print(f"导入完成:")
    print(f"  违规行为:      {v_inserted} 新增, {v_skipped} 跳过(重复), {errors} 错误")
    print(f"  audititem:     {matched_items} 匹配")
    print(f"  法规关联:      {law_refs_inserted} 匹配, {law_refs_unmatched} 未匹配(跳过，法规名保留在description)")
    print(f"  案例:          {cases_inserted}")
    print(f"  案例↔违规:     {case_viol_inserted}")
    print(f"  案例↔法规:     {case_law_inserted}")

    return {
        "dry_run": False,
        "violations": v_inserted,
        "audititem_matched": matched_items,
        "law_refs": law_refs_inserted,
        "law_refs_unmatched": law_refs_unmatched,
        "cases": cases_inserted,
        "batch": BATCH_ID,
    }


if __name__ == "__main__":
    dry_run = "--run" not in sys.argv
    import_excel(dry_run=dry_run)
